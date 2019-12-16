'''
graph.py (Excel Command Line Tool)
Sydney Fowler and Matthew Hileman
15 December 2019
Description: Allows the user to select a column of data for the x axis and a column of data for the y axis and plots
the data using matplotlib.
'''

# ================ REFERENCES ================
# OPENPYXL (needed import)

# ================ IMPORTS ================
# Custom
from excel_funcs import get_directory
import menus

# Exterior
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter

# ================== SETUP ===================
def menu_header():
    # Menus object
    duplicate_main_menu = menus.Menu("graph", menus.GRAPH_LIST, menus.GRAPH_ROUTE)
    duplicate_main_menu.print_menu_message()
    duplicate_main_menu.display_shift_menu()

def init():
    # Get workbook
    wb_path = get_directory([".xlsx"], "Type path of your excel file (.xlsx): ")
    wb = openpyxl.load_workbook(wb_path)
    sheets = wb.sheetnames  # Edited depreciated function: "wb.get_sheet_names()"

    # Select sheet
    sheet_menu = menus.Value_Menu("cleanup", sheets, sheets)
    user_selection = sheet_menu.display_shift_menu()
    sheet = wb.get_sheet_by_name(user_selection)
    wb.active = sheet

    # Select x-axis and y-axis
    headers = []
    for cell in sheet[1]:
        headers.append(str(cell.value))
    header_menu = menus.Value_Menu("cleanup", headers, headers)
    x_axis = header_menu.display_shift_menu()
    x = np.empty(sheet.max_row - 1)
    rows = list(sheet.iter_rows(min_row=2,
                                max_row=sheet.max_row,
                                min_col=headers.index(x_axis) + 1,
                                max_col=headers.index(x_axis) + 1))
    for i in range(sheet.max_row - 1):
        try:
            x[i] = rows[i][0].value
        except Exception:
            print("ERROR: Inconsistent data type for x on " + str(rows[i][0].value)
                  + ". All data in x must be of same type.")
            return
    y_axis = header_menu.display_shift_menu()
    y = np.empty(sheet.max_row - 1)
    rows = list(sheet.iter_rows(min_row=2,
                                max_row=sheet.max_row,
                                min_col=headers.index(y_axis) + 1,
                                max_col=headers.index(y_axis) + 1))
    for i in range(sheet.max_row - 1):
        try:
            y[i] = rows[i][0].value
        except Exception:
            print("ERROR: Inconsistent data type for y on " + str(rows[i][0].value)
                  + ". All data in x must be of same type.")
            return
    plt.plot(x, y, "oy")
    plt.show()


# For test purposes, will execute header if being run as main
if __name__ == '__main__':
    menu_header()
