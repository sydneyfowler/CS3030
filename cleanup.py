'''
cleanup.py (Excel Command Line Tool)
Sydney Fowler and Matthew Hileman
15 December 2019
Description: Allows the user to select a set of cleanup rules for each column in their file and applies said cleanup
to a new version of the file.
'''

# ================ REFERENCES ================
# OPENPYXL (needed import)

# ================ IMPORTS ================
# System
import os
from datetime import datetime

# Custom
from custom_regular_expressions import strip_none_digits
from custom_regular_expressions import remove_special_characters
from custom_regular_expressions import phone_regex
from custom_regular_expressions import email_regex
from custom_regular_expressions import zip_regex
from custom_regular_expressions import yyyy_mm_dd
from custom_regular_expressions import mm_dd_yyyy
from custom_regular_expressions import month_word
from custom_regular_expressions import web_address_regex

import custom_dictionaries
from excel_funcs import get_directory
from excel_funcs import save_file
import menus

# Exterior
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ================== SETUP ===================
# CONSTANTS
CLEANUP_OPTIONS_LIST = ["Cleanup Phone Numbers", "Cleanup Email Addresses", "Cleanup States", "Cleanup Zip Codes",
                        "Cleanup Dates", "Cleanup Web Address", "Produce List of Unique Entries",
                        "Check Entries Against List", "Truncate to Character Limit", "Check Data Type",
                        "No Cleaning", "Finish Sheet"]
DATA_TYPE_LIST = ["Whole Number", "Decimal Value", "Currency", "Text String", "Date", "Not Specified"]

NO_CLEANING = CLEANUP_OPTIONS_LIST.index("No Cleaning")
BREAK_SHEET = CLEANUP_OPTIONS_LIST.index("Finish Sheet")


def menu_header():
    # Print Main Cleanup Menu
    cleanup_main_menu = menus.Menu("cleanup", menus.CLEANUP_MENU_LIST, menus.CLEANUP_MENU_ROUTE)
    cleanup_main_menu.print_menu_message()
    cleanup_main_menu.display_shift_menu()


def init():
    # Get workbook
    wb_path = get_directory([".xlsx"], "Type path of your excel file (.xlsx): ")
    wb = openpyxl.load_workbook(wb_path)
    sheets = wb.get_sheet_names()

    # Initialize 2D dictionary representing each sheet and its column headers
    sheet_header_lookup = {}
    for sheet_name in sheets:
        sheet_header_lookup.setdefault(sheet_name, {})
        sheet = wb.get_sheet_by_name(sheet_name)
        for cell in sheet[1]:
            sheet_header_lookup[sheet_name].setdefault(cell.value, None)

    # Get user selections
    for sheet_name in sheets:
        # Check if user wants to process this sheet
        process_sheet = input("Would you like to clean sheet " + sheet_name + "? (y/n) ")
        if process_sheet not in ("yes", "Yes", "Y", "y"):
            continue

        # Get user selections for each header in sheet
        for header in sheet_header_lookup[sheet_name]:
            print_menu(sheet_name, header, CLEANUP_OPTIONS_LIST)
            user_selection = get_user_selection(CLEANUP_OPTIONS_LIST)
            if user_selection == BREAK_SHEET:       # Check if user wants to break out of sheet
                break
            elif user_selection == NO_CLEANING:     # Check if user wants to skip this column
                continue
            else:
                sheet_header_lookup[sheet_name][header] = user_selection

    # Process data
    for sheet_name in sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        for col in range(1, sheet.max_column + 1):
            process_number = sheet_header_lookup[sheet_name][sheet.cell(row=1, column=col).value]
            if process_number is not None:
                col_letter = get_column_letter(col)
                process_column(wb, sheet[col_letter], int(process_number))

    # Save to a new copy of the workbook
    new_file = wb_path[:len(wb_path) - 5] + "_EDITED.xlsx"
    wb.save(new_file)


def print_menu(sheet_name, header, l):
    print('-' * 40)
    print("SHEET: " + str(sheet_name))
    print("HEADER: " + str(header))
    print("Select an option (0-" + str(len(l) - 1) + ")")
    print('-' * 40)
    # Prints each item in list
    for item in l:
        print("(" + str(l.index(item)) + ") " + item)


def get_user_selection(l):
    # Error checking loop - input is an integer and is a valid menu item
    while (True):
        print('-' * 40)
        print("Choice: ", end='')

        # Initialize choice
        user_choice = input()

        # Error handling: makes sure input is integer, stores interger
        try:
            user_choice = int(user_choice)
        # If input is not an integer, display error, has user try again.
        except ValueError:
            # Error Message
            print()
            print(str(user_choice) + " is not a valid input (NOT AN INT)")
            print("Choose a numeric value from the options above between (0-" + str(len(l) - 1)
                  + ").")
            continue

        # Error handling: makes sure the user's choice is a valid menu option
        if (user_choice < 0) or (user_choice >= len(l)):
            # Error Message
            print()
            print(str(user_choice) + " is not a valid input.")
            print("Choose a numeric value from the options above between (0-" + str(len(l) - 1)
                  + ").")
            continue

        # If input is valid, return the input value, break from error loop
        else:
            return user_choice


def process_column(wb, range, process_number):
    if process_number == CLEANUP_OPTIONS_LIST.index("Cleanup Phone Numbers"):
        clean_phone_number(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Cleanup Email Addresses"):
        clean_email_address(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Cleanup States"):
        clean_states(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Cleanup Zip Codes"):
        clean_zip_codes(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Cleanup Dates"):
        clean_dates(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Cleanup Web Address"):
        clean_web_addresses(wb_range)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Produce List of Unique Entries"):
        get_unique_entries(wb_range, wb)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Check Entries Against List"):
        user_list = get_list(wb_range[0].value)
        check_entries_against_list(wb_range, user_list)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Truncate to Character Limit"):
        limit = get_limit(wb_range[0].value)
        check_character_limit(wb_range, limit)
    elif process_number == CLEANUP_OPTIONS_LIST.index("Check Data Type"):
        data_type = get_data_type(wb_range[0].value)
        check_data_type(range, data_type)


def get_list(header):
    while (True):  # Loop until you get a valid text file
        print('-' * 40)
        list_file_path = input("Type path of the text file containing the list you would like " + header
                               + " checked against: ")
        list_file_path = os.path.abspath(list_file_path)
        if os.path.exists(list_file_path):
            if list_file_path[-4:] == ".txt":
                try:
                    user_list_file = open(list_file_path)
                    user_list = user_list_file.read().splitlines()
                    user_list_file.close()
                    break
                except Exception:
                    print("ERROR: Unable to open file.")
                    print()
            else:
                print("ERROR: Must be a .txt file.")
                print()
        else:
            print("ERROR: Invalid file path.")
            print()
    return user_list


def get_limit(header):
    while (True):  # Loop until you get a valid text file
        print('-' * 40)
        limit = input("Enter the character limit you would like used for " + header + ": ")
        try:
            limit = int(limit)
            break
        except Exception:
            print("ERROR: Must be a whole number.")
            print()
    return limit


def get_data_type(header):
    print('-' * 40)
    print("Select the data type option (0-" + str(len(DATA_TYPE_LIST) - 1) + ") you would like used for " + header)
    print('-' * 40)
    # Prints each item in list
    for item in DATA_TYPE_LIST:
        print("(" + str(DATA_TYPE_LIST.index(item)) + ") " + item)
    return get_user_selection(DATA_TYPE_LIST)


def clean_phone_number(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regex
        if phone_regex.search(str(cell.value)):
            match = phone_regex.search(str(cell.value))
            phone_number = ""
            if match.group('area_code'):
                area_code = strip_none_digits.search(match.group('area_code'))
                phone_number += "(" + area_code.group(0) + ") " + match.group('three_digits') + "-" \
                                + match.group('four_digits')
            else:
                phone_number += match.group('three_digits') + "-" + match.group('four_digits')
            if match.group('ext'):
                phone_number += 'x' + match.group('ext')
            cell.value = phone_number
        else:
            cell.value = ""


def clean_email_address(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regex
        if email_regex.search(str(cell.value)):
            match = email_regex.search(str(cell.value))
            cell.value = match.group(1)
        else:
            cell.value = ""


def clean_states(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Look for cell.value in states_lookup dictionary
        state = ((remove_special_characters.sub("", str(cell.value))).upper()).strip()
        if state in custom_dictionaries.states_lookup.keys():
            cell.value = custom_dictionaries.states_lookup[state]
        else:
            cell.value = ""


def clean_zip_codes(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regex
        if zip_regex.search(str(cell.value)):
            match = zip_regex.search(str(cell.value))
            # Pad five digits if needed
            number_of_zeros = 5 - len(str(match.group('five_digits')))
            zip_code = ("0" * number_of_zeros) + match.group('five_digits')
            if match.group('four_digits'):
                zip_code += "-" + match.group('four_digits')
            cell.value = zip_code
        else:
            cell.value = ""


def clean_dates(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regexes
        if yyyy_mm_dd.search(str(cell.value)):
            match = yyyy_mm_dd.search(str(cell.value))
            day = "0" * (2 - len(match.group('day'))) + match.group('day')
            month = "0" * (2 - len(match.group('month'))) + match.group('month')
            cell.value = match.group('year') + "-" + month + "-" + day

        elif mm_dd_yyyy.search(str(cell.value)):
            match = mm_dd_yyyy.search(str(cell.value))
            day = "0" * (2 - len(match.group('day'))) + match.group('day')
            month = "0" * (2 - len(match.group('month'))) + match.group('month')
            year = ""
            if int(match.group('year')) < 100:
                if int("20" + str(match.group('year'))) < (datetime.now()).year:
                    year += "20" + str(match.group('year'))
                else:
                    year += "19" + str(match.group('year'))
            else:
                year += str(match.group('year'))
            cell.value = year + "-" + month + "-" + day

        elif month_word.search(str(cell.value).upper()):
            match = month_word.search(str(cell.value).upper())
            day = "0" * (2 - len(match.group('day'))) + match.group('day')
            year = ""
            if int(match.group('year')) < 100:
                if int("20" + str(match.group('year'))) <= (datetime.now()).year:
                    year += "20" + str(match.group('year'))
                else:
                    year += "19" + str(match.group('year'))
            else:
                year += str(match.group('year'))
            month = custom_dictionaries.month_lookup[(str(match.group('month')).upper())[:3]]
            cell.value = year + "-" + month + "-" + day

        else:
            cell.value = ""


def clean_web_addresses(range):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regex
        if web_address_regex.search(str(cell.value)):
            match = web_address_regex.search(str(cell.value))
            cell.value = match.group(1)
        else:
            cell.value = ""


def get_unique_entries(wb_range, wb):
    entries = []
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check entries for cell.value (case-insensitive), if not there, add to entries
        if cell.value:
            s = (str(cell.value)).upper()
            if s not in entries:
                entries.append((str(cell.value)).upper())

    # Create new sheet in Excel workbook for the entries
    wb.create_sheet(title=wb_range[0].value)
    sheet = wb.get_sheet_by_name(wb_range[0].value)
    my_font = Font(bold=True)
    sheet['A1'].font = my_font
    sheet['A1'].value = wb_range[0].value
    for i in range(0, len(entries)):
        sheet.cell(row=i+2, column=1).value = entries[i].title()


def check_entries_against_list(range, l):
    # Convert list to uppercase so the check is case-insensitive
    for item in l:
        item = item.upper()

    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Look for cell.value in user_list, if not there, remove the entry in the Excel File
        if (str(cell.value)).upper() not in l:
            cell.value = ""


def check_character_limit(range, limit):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check if cell.value is within character limit, if not, truncate
        if len(str(cell.value)) > limit:
            cell.value = (str(cell.value))[:limit]


def check_data_type(range, t):
    # Clean column
    for cell in range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Check against regex
        if t == 0:                              # Whole Number
            try:
                int(cell.value)
                cell.value = int(cell.value)
            except Exception:
                cell.value = ""
        elif t == 1:                            # Decimal Value
            try:
                float(cell.value)
                cell.value = float(cell.value)
            except Exception:
                cell.value = ""
        elif t == 2:                            # Currency
            try:
                float(cell.value)
                cell.value = "$" + "{:.2f}".format(round(float(cell.value), 2))
            except Exception:
                cell.value = ""
        elif t == 3:                            # Text String
            continue
        elif t == 4:                            # Datetime Stamp
            s = str(cell.value)
            if yyyy_mm_dd.search(s):
                match = yyyy_mm_dd.search(s)
                cell.value = match.group(1)
            elif mm_dd_yyyy.search(s):
                match = mm_dd_yyyy.search(s)
                cell.value = match.group(1)
            elif month_word.search(s.upper()):
                match = month_word.search(s.upper())
                cell.value = (match.group(1)).title()
            else:
                cell.value = ""
        elif t == 5:                            # Not Specified
            continue


# For test purposes, will execute header if being run as main
if __name__ == '__main__':
    menu_header()
