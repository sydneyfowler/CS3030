'''
duplicate_removal.py (Excel Command Line Tool)
Sydney Fowler and Matthew Hileman
15 December 2019
Description: Removes duplicates from rows, columns, or entire sheets.
'''

<<<<<<< HEAD
# ================ REFERENCES ================
# OPENPYXL - needed import

=======
>>>>>>> master
# ================ IMPORTS ================
# System
import os
import sys

# Custom
from excel_funcs import get_directory
from excel_funcs import save_file
from excel_funcs import get_sheet
import menus
import main

# Exterior
import openpyxl
<<<<<<< HEAD
=======

# ================ REFERENCES ================
# OPENPYXL - needed import
>>>>>>> master

# ================== SETUP ===================
def menu_header():

    # Print Import Message Above
    duplicate_main_menu = menus.Menu("duplicate_removal", menus.DUPLICATE_MENU_LIST, menus.DUPLICATE_MENU_ROUTE)
    duplicate_main_menu.print_menu_message()
    duplicate_main_menu.display_shift_menu()

def init():

    # Get user input for excel file
    wb_path = get_directory([".xlsx"], "Type path of your excel file (.xlsx): ")
    wb = openpyxl.load_workbook(wb_path)

    # Gets input from user if they want to remove duplicates from all sheets
    if (len(wb.sheetnames) > 1):      # Edited depreciated function: "wb.get_sheet_names()"
        user_choice = input ("Would you like to remove duplicates from all sheets (Y) or one sheet (N)?: ")
        if user_choice in ("yes", "Yes", "Y", "y"):
            remove_duplicate_all_sheet(wb, wb_path)
        else:
            single_sheet(wb, wb_path)
    else:
        # If only 1 sheet, chooses single sheet option
        single_sheet(wb, wb_path)

    # Loop back to top menu
    menu_header()

# Removes duplicates from all sheets
def remove_duplicate_all_sheet(wb, wb_path):
    print("Removing duplicate rows from all sheets...")

    sheets = wb.sheetnames      # Edited depreciated function: "wb.get_sheet_names()"

    # Iterate over each sheet
    for item in sheets:

        # Create sheet object
        sheet = wb.get_sheet_by_name(item)

        # Initialize rows list, duplicate list, row index
        rows_list = []
        dup_count = 0
        row_num = 1
        duplicate_flag = 0

        # Iterate over each row
        for row in sheet.iter_rows(min_row = row_num):

            # Store current row as a list
            current_row = []
            for cellObj in row:
                current_row.append(cellObj.value)

            # If row is identical, add to duplicates list
            if current_row in rows_list:
                sheet.delete_rows(row_num, 1)
                dup_count += 1
            # Else, add to list of already read rows
            else:
                rows_list.append(current_row)
                row_num += 1

        # Print how many rows were deleted in what sheets
        if (dup_count > 0):
            duplicate_flag = 1
            print("Removed " + str(dup_count) + " duplicate rows from " + item)

    # Return for change_file_flag in init()
    if not duplicate_flag:
        print("Tool could not find any duplicate rows to remove!")
        input("Press enter to continue...")
    else:
        save_file(wb, wb_path, ".xlsx")

# If user chooses to remove duplicates from a single sheet
def single_sheet(wb, wb_path):

    # CONSTANTS
    REMOVAL_MENU_LIST = ["Remove duplicates based on row", "Remove duplicates based on column",
                        "Go back to top menu", "Exit duplicate removal tool"]
    REMOVAL_MENU_ROUTE = ["remove_duplicate_rows", "remove_duplicate_cols",
                        "exit_tool", "go_main"]

    # Removal menu, select to remove by a certain row or column
    removal_menu = menus.Function_Menu("duplicate_removal", REMOVAL_MENU_LIST, REMOVAL_MENU_ROUTE)
    removal_menu.display_shift_menu(wb, wb_path)

# Removes rows based on a column
def remove_duplicate_rows(args):

    wb = args[0]
    wb_path = args[1]
    sheet = get_sheet(wb)

    # Select row to base removal on
    row_index = range(0, sheet.max_column - 1)
    row_menu = menus.Value_Menu("duplicate_removal", row_index, row_index)

    print()
    print(" Choose a column to base duplicate removal on ".center(70, "="))
    print('-' * 70)

    # Which col to compare
    removal_index = row_menu.display_shift_menu()

    # Initialize rows list, duplicate list, row index
    rows_list = []
    dup_count = 0
    row_num = 1
    duplicate_flag = 0

    # Iterate over each row
    for row in sheet.iter_rows(min_row = row_num):

        # Store current row's repsective col value
        current_row = row[removal_index].value

        # If row is identical, add to duplicates list
        if current_row in rows_list:
            sheet.delete_rows(row_num, 1)
            dup_count += 1
        # Else, add to list of already read rows
        else:
            rows_list.append(current_row)
            row_num += 1

    # Print how many rows were deleted in what sheets
    if (dup_count > 0):
        duplicate_flag = 1
        print("Removed " + str(dup_count) + " duplicate rows from list.")

    # Return for change_file_flag in init()
    if not duplicate_flag:
        print("Tool could not find any duplicate rows to remove!")
        input("Press enter to continue...")
    else:
        save_file(wb, wb_path, ".xlsx")


# Removes columns based on a row
def remove_duplicate_cols(args):

        wb = args[0]
        wb_path = args[1]
        sheet = get_sheet(wb)

        # Select row to base removal on
        col_index = range(0, sheet.max_row)
        col_name = range(1, sheet.max_row + 1)
        col_menu = menus.Value_Menu("duplicate_removal", col_name, col_index)

        print()
        print(" Choose a column to base duplicate removal on ".center(70, "="))
        print('-' * 70)

        # Which col to compare
        removal_index = col_menu.display_shift_menu()

        # Initialize col list, duplicate list, col index
        cols_list = []
        dup_count = 0
        col_num = 1
        duplicate_flag = 0

        # Iterate over each col
        for col in sheet.iter_cols(min_col = col_num):

            # Store current col repsective row value
            current_col = col[removal_index].value

            # If row is identical, add to duplicates list
            if current_col in cols_list:
                sheet.delete_cols(col_num, 1)
                dup_count += 1
            # Else, add to list of already read rows
            else:
                cols_list.append(current_col)
                col_num += 1

        # Print how many rows were deleted in what sheets
        if (dup_count > 0):
            duplicate_flag = 1
            print("Removed " + str(dup_count) + " duplicate rows from list.")

        # Return for change_file_flag in init()
        if not duplicate_flag:
            print("Tool could not find any duplicate rows to remove!")
            input("Press enter to continue...")
        else:
            save_file(wb, wb_path, ".xlsx")

# Returns user to top menu of this tool
def exit_tool(wb):
    menu_header()

# Returns user to main.py
def go_main(wb):
    main.menu_header()

# For test purposes, will execute header if being run as main
if __name__ == '__main__':
    menu_header()
