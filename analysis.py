'''
analysis.py (Excel Command Line Tool)
Sydney Fowler and Matthew Hileman
15 December 2019
Description: Allows the user to select from a set of analysis options for each column in their file. For each sheet
selected, a new analysis sheet is created with the results from each analyzed column.
'''

# ================ REFERENCES ================
# OPENPYXL (needed import)

# ================ IMPORTS ================
# Custom
from excel_funcs import get_directory
from excel_funcs import save_file
import menus

# Exterior
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# ================== SETUP ===================
# CONSTANTS
ANALYSIS_OPTIONS_LIST = ["Sum", "Count", "Max", "Min", "Check Unique", "Average", "No Analysis", "Finish Sheet"]
NO_ANALYSIS = "No Analysis"
BREAK_SHEET = "Finish Sheet"


def menu_header():
    # Print Main Analysis Menu
    cleanup_main_menu = menus.Menu("analysis", menus.ANALYSIS_MENU_LIST, menus.ANALYSIS_MENU_ROUTE)
    cleanup_main_menu.print_menu_message()
    cleanup_main_menu.display_shift_menu()


def init():
    # Get workbook
    wb_path = get_directory([".xlsx"], "Type path of your excel file (.xlsx): ")
    wb = openpyxl.load_workbook(wb_path)
    sheets = wb.sheetnames      # Edited depreciated function: "wb.get_sheet_names()"

    # Initialize 2D dictionary representing each sheet and its column headers
    sheet_header_lookup = {}
    for sheet_name in sheets:
        sheet_header_lookup.setdefault(sheet_name, {})
        sheet = wb.get_sheet_by_name(sheet_name)
        for cell in sheet[1]:
            sheet_header_lookup[sheet_name].setdefault(cell.value, None)

    # Get user selections for each sheet
    for sheet_name in sheets:

        # Check if user wants to process this sheet
        analyze_sheet = input("Would you like to analyze sheet " + sheet_name + "? (y/n) ")
        if analyze_sheet not in ("yes", "Yes", "Y", "y"):
            continue

        # Menu object used below
        cleanup_menu = menus.Value_Menu("cleanup", ANALYSIS_OPTIONS_LIST, ANALYSIS_OPTIONS_LIST)

        # Get user selections for each header in sheet
        for header in sheet_header_lookup[sheet_name]:
            print('-' * 40)
            print("SHEET: " + str(sheet_name))
            print("HEADER: " + str(header))
            user_selection = cleanup_menu.display_shift_menu()

            if user_selection == BREAK_SHEET:       # Check if user wants to break out of sheet
                break
            elif user_selection == NO_ANALYSIS:     # Check if user wants to skip this column
                continue
            else:
                sheet_header_lookup[sheet_name][header] = user_selection

    # Analyze data
    change_file_flag = 0
    for sheet_name in sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        for col in range(1, sheet.max_column + 1):
            analysis_number = sheet_header_lookup[sheet_name][sheet.cell(row=1, column=col).value]
            if analysis_number is not None:
                change_file_flag = 1
                col_letter = get_column_letter(col)
                perform_analysis(wb, sheet_name, sheet[col_letter], int(analysis_number))

    # Save new file
    if change_file_flag:
        save_file(wb, wb_path, ".xlsx")
    else:
        print()
        print("File not changed, no need to save new vesion.")
        input("Press enter to continue...")

    # Loop back to top menu
    menu_header()


def perform_analysis(wb, sheet_name, wb_range, analysis_number):
    if analysis_number == ANALYSIS_OPTIONS_LIST.index("Sum"):
        op = "SUM"
        ret_val = perform_sum(sheet_name, wb_range)
    elif analysis_number == ANALYSIS_OPTIONS_LIST.index("Count"):
        op = "COUNT"
        ret_val = perform_count(sheet_name, wb_range)
    elif analysis_number == ANALYSIS_OPTIONS_LIST.index("Max"):
        op = "MAX"
        ret_val = perform_max(sheet_name, wb_range)
    elif analysis_number == ANALYSIS_OPTIONS_LIST.index("Min"):
        op = "MIN"
        ret_val = perform_min(sheet_name, wb_range)
    elif analysis_number == ANALYSIS_OPTIONS_LIST.index("Check Unique"):
        op = "UNIQUE"
        ret_val = perform_unique(sheet_name, wb_range)
    elif analysis_number == ANALYSIS_OPTIONS_LIST.index("Average"):
        op = "AVERAGE"
        ret_val = perform_average(sheet_name, wb_range)

    # If analysis sheet is not created, create it
    if sheet_name.upper() + " ANALYSIS" not in wb.get_sheet_names():
        wb.create_sheet(title=sheet_name.upper() + " ANALYSIS")
        sheet = wb.get_sheet_by_name(sheet_name.upper() + " ANALYSIS")
        my_font = Font(bold=True)
        sheet.merge_cells('A1:C1')
        sheet['A1'].font = my_font
        sheet['A1'].value = "ANALYSIS SUMMARY"
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A2'].font = my_font
        sheet['A2'].value = "Column"
        sheet['B2'].font = my_font
        sheet['B2'].value = "Operation"
        sheet['C2'].font = my_font
        sheet['C2'].value = "Result"

    sheet = wb.get_sheet_by_name(sheet_name.upper() + " ANALYSIS")
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = wb_range[0].value
    sheet.cell(row=new_row, column=2).value = op
    sheet.cell(row=new_row, column=3).value = ret_val


def perform_sum(sheet_name, wb_range):
    ret_sum = 0
    non_float_flag = False

    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        try:
            ret_sum += float(cell.value)
        except Exception:
            non_float_flag = True

    # Print warning message if some of the cell values could not be summed
    if non_float_flag:
        print("WARNING: Some cells in " + sheet_name + ":" + wb_range[0].value + " were not numerical values.")

    if ret_sum == 0:
        return("N/A")
    else:
        return str(ret_sum)


def perform_count(sheet_name, wb_range):
    ret_count = 0

    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        if cell.value:
            ret_count += 1

    return str(ret_count)


def perform_max(sheet_name, wb_range):
    ret_max = "N/A"
    non_float_flag = False

    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        try:
            num = float(cell.value)
            if ret_min == "N/A":
                ret_min = num
            elif num > ret_max:
                ret_max = num
        except Exception:
            non_float_flag = True

    # Print warning message if some of the cell values could not be summed
    if non_float_flag:
        print("WARNING: Some cells in " + sheet_name + ":" + wb_range[0].value + " were not numerical values.")

    return str(ret_max)


def perform_min(sheet_name, wb_range):
    ret_min = "N/A"
    non_float_flag = False

    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        try:
            num = float(cell.value)
            if ret_min == "N/A":
                ret_min = num
            elif num < ret_min:
                ret_min = num
        except Exception:
            non_float_flag = True

    # Print warning message if some of the cell values could not be summed
    if non_float_flag:
        print("WARNING: Some cells in " + sheet_name + ":" + wb_range[0].value + " were not numerical values.")

    return str(ret_min)


def perform_unique(sheet_name, wb_range):
    entries = []
    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        if cell.value:
            if cell.value.lower() in entries:
                return "False"
            entries.append(cell.value.lower())

    return "True"


def perform_average(sheet_name, wb_range):
    init_sum = 0
    init_count = 0
    non_float_flag = False

    # Analyze column
    for cell in wb_range:
        # Skip Header Row
        if cell.row == 1:
            continue
        # Perform op
        try:
            init_sum += float(cell.value)
            init_count += 1
        except Exception:
            non_float_flag = True

    # Print warning message if some of the cell values could not be summed
    if non_float_flag:
        print("WARNING: Some cells in " + sheet_name + ":" + wb_range[0].value + " were not numerical values.")

    if init_count == 0:
        return 0
    else:
        return str(round(init_sum / init_count, 2))


# For test purposes, will execute header if being run as main
if __name__ == '__main__':
    menu_header()
